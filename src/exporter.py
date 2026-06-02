from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

NAVY = "1F3A5F"  
BLUE = "2E86C1"
LIGHT = "D6EAF8"
STRIPE = "F2F7FB"
GREEN = "1E8449"
RED = "C0392B"
GREY = "566573"

WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
SUB_FONT = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=NAVY)
LABEL_FONT = Font(name="Calibri", size=11, color=GREY)
VALUE_FONT = Font(name="Calibri", size=11, bold=True, color="1A1A1A")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
BLUE_FILL = PatternFill("solid", fgColor=BLUE)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
STRIPE_FILL = PatternFill("solid", fgColor=STRIPE)

THIN = Side(style="thin", color="D5DBDB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

MONEY = '#,##0;[Red](#,##0)'
MONEY2 = '#,##0.00;[Red](#,##0.00)'
PCT = '0.00%'

class ExcelExporter:
    """Render StockMetrics + DCF analysis into a styled Excel workbook."""

    def __init__(self, data: dict, dcf, output_dir: str = "data"):
        """
        Parameters:
        data (dict): The dict returned by StockMetrics.get_data().
        dcf (DCF): A calculated DCF instance for the same ticker.
        output_dir (str): Folder the workbook is written into.
        """
        self.data = data
        self.dcf = dcf
        self.ticker = data["ticker"]
        self.output_dir = Path(output_dir)
        self.wb = Workbook()

    # -- public ---------------------------------------------------------------
    def export(self, filename: str | None = None) -> Path:
        """Build every sheet and save the workbook, returning its path."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        if filename is None:
            stamp = datetime.now().strftime("%Y%m%d")
            filename = f"{self.ticker}_valuation_{stamp}.xlsx"
        path = self.output_dir / filename

        self.wb.remove(self.wb.active) # drop default empty sheet
        self._build_summary()
        self._build_dcf()
        self._build_sensitivity()
        self._build_price_data()

        self.wb.save(path)
        return path

    # -- sheets ---------------------------------------------------------------
    def _build_summary(self) -> None:
        ws = self.wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 3

        self._banner(ws, f"{self.ticker} — Valuation Summary", f"Generated {datetime.now():%B %d, %Y}", first="B", last="C")

        pe = self.data["P/E Ratio"]
        vol = self.data["Volatility"]
        s = self.dcf.summary()
        price = self.dcf.info.get("currentPrice") or pe.get("Price")
        intrinsic = s["intrinsic_value"]
        mos = s["margin_of_safety"]
        verdict = "UNDERVALUED" if mos > 0 else "OVERVALUED"

        row = 4
        row = self._section(ws, row, "Market Snapshot")
        row = self._kv(ws, row, "Current Price", price, MONEY2, prefix="$")
        row = self._kv(ws, row, "P/E (TTM)", pe.get("P/E (TTM)"))
        row = self._kv(ws, row, "Forward P/E", pe.get("Forward P/E"))
        row = self._kv(ws, row, "EPS (TTM)", pe.get("EPS (TTM)"), prefix="$")
        row = self._kv(ws, row, "Market Cap", pe.get("Market Cap"))
        row = self._kv(ws, row, "Annualized Volatility", vol.get("Annualized Volatility Pct"))

        row += 1
        row = self._section(ws, row, "Intrinsic Valuation (DCF)")
        row = self._kv(ws, row, "Enterprise Value", s["enterprise_value"], MONEY, prefix="$")
        row = self._kv(ws, row, "Equity Value", s["equity_value"], MONEY, prefix="$")
        row = self._kv(ws, row, "Intrinsic Value / Share", intrinsic, MONEY2, prefix="$", emphasize=True)
        row = self._kv(ws, row, "Margin of Safety", mos, PCT, color=GREEN if mos > 0 else RED, emphasize=True)

        row += 1
        cell = ws.cell(row=row, column=2, value=verdict)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        cell.alignment = CENTER
        cell.fill = PatternFill("solid", fgColor=GREEN if mos > 0 else RED)
        ws.row_dimensions[row].height = 26

    def _build_dcf(self) -> None:
        ws = self.wb.create_sheet("DCF Detail")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 28

        # Size value columns to fit the largest figure they'll hold
        s = self.dcf.summary()
        figures = [
            self.dcf.base_fcf, s["terminal_value"], s["pv_terminal_value"],
            s["enterprise_value"], sum(self.dcf.pv_fcfs), s["total_debt"],
            s["cash_and_st_investments"], s["lt_investments"], s["equity_value"],
            *self.dcf.projected_fcfs, *self.dcf.pv_fcfs,
        ]

        width = max(16, max(len(f"${v:,.0f}") for v in figures) + 3)
        for col in range(3, 3 + self.dcf.projection_years):
            ws.column_dimensions[get_column_letter(col)].width = width

        self._banner(ws, f"{self.ticker} - DCF Projection", "All figures in absolute currency units", first="B", last=get_column_letter(2 + self.dcf.projection_years))

        row = 4
        row = self._section(ws, row, "Assumptions", last_col=2 + self.dcf.projection_years)
        row = self._kv(ws, row, "Base FCFF (3yr avg)", self.dcf.base_fcf, MONEY, prefix="$")
        row = self._kv(ws, row, "Tax Rate", self.dcf.tax_rate, PCT)
        row = self._kv(ws, row, "Discount Rate (WACC)", self.dcf.discount_rate, PCT)
        row = self._kv(ws, row, "Growth Rate", self.dcf.growth_rate, PCT)
        row = self._kv(ws, row, "Terminal Growth", self.dcf.terminal_growth_rate, PCT)
        row = self._kv(ws, row, "Projection Years", self.dcf.projection_years, "0")

        row += 1

        years = [f"Year {i}" for i in range(1, self.dcf.projection_years + 1)]
        self._table_header(ws, row, ["Cash Flow"] + years, start_col=2)
        row += 1
        row = self._table_row(ws, row, "Projected FCF", self.dcf.projected_fcfs, MONEY, stripe=False)
        row = self._table_row(ws, row, "Present Value of FCF", self.dcf.pv_fcfs, MONEY, stripe=True)

        row += 1
        row = self._section(ws, row, "Terminal & Totals", last_col=2 + self.dcf.projection_years)
        row = self._kv(ws, row, "Terminal Value", s["terminal_value"], MONEY, prefix="$")
        row = self._kv(ws, row, "PV of Terminal Value", s["pv_terminal_value"], MONEY, prefix="$")
        row = self._kv(ws, row, "Sum PV of FCFs", sum(self.dcf.pv_fcfs), MONEY, prefix="$")
        row = self._kv(ws, row, "Enterprise Value", s["enterprise_value"], MONEY, prefix="$", emphasize=True)

        # Equity bridge: EV - debt + cash + short-term + non-operating long-term
        # investments -> equity value -> per share. Subtractions shown negative.
        row += 1
        last_col = 2 + self.dcf.projection_years
        row = self._section(ws, row, "Equity Bridge", last_col=last_col)
        row = self._kv(ws, row, "Enterprise Value", s["enterprise_value"], MONEY, prefix="$")
        row = self._kv(ws, row, "Less: Total Debt", -s["total_debt"], MONEY, prefix="$", color=RED)
        row = self._kv(ws, row, "Add: Cash & ST Investments", s["cash_and_st_investments"], MONEY, prefix="$", color=GREEN)
        row = self._kv(ws, row, "Add: LT Investments (non-op)", s["lt_investments"], MONEY, prefix="$", color=GREEN)
        row = self._kv(ws, row, "Equity Value", s["equity_value"], MONEY, prefix="$", emphasize=True)
        row = self._kv(ws, row, "Shares Outstanding", self.dcf.shares_outstanding, "#,##0")
        row = self._kv(ws, row, "Intrinsic Value / Share", s["intrinsic_value"], MONEY2, prefix="$", emphasize=True)
        
        note = ws.cell(row=row, column=2, value=f"Net debt source: {s['net_debt_source']}")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
        note.font = Font(name="Calibri", size=9, italic=True, color=GREY)

    def _build_sensitivity(self) -> None:
        ws = self.wb.create_sheet("Sensitivity")
        ws.sheet_view.showGridLines = False
        table = self.dcf.sensitivity_table()

        self._banner(ws, f"{self.ticker} — Sensitivity Analysis", "Intrinsic value / share — rows: discount rate, columns: growth rate", first="B", last=get_column_letter(2 + len(table.columns)))

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 18
        for i in range(len(table.columns)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 13

        start_row = 4
        corner = ws.cell(row=start_row, column=2, value="Disc \\ Growth")
        corner.font = WHITE_BOLD
        corner.fill = NAVY_FILL
        corner.alignment = CENTER
        corner.border = BORDER
        for j, col in enumerate(table.columns):
            c = ws.cell(row=start_row, column=3 + j, value=col)
            c.font = WHITE_BOLD
            c.fill = BLUE_FILL
            c.alignment = CENTER
            c.border = BORDER

        first_data_row = start_row + 1
        for i, (idx, series) in enumerate(table.iterrows()):
            r = first_data_row + i
            h = ws.cell(row=r, column=2, value=idx)
            h.font = WHITE_BOLD
            h.fill = BLUE_FILL
            h.alignment = CENTER
            h.border = BORDER
            for j, val in enumerate(series):
                c = ws.cell(row=r, column=3 + j,
                            value=None if pd.isna(val) else float(val))
                c.number_format = MONEY2
                c.alignment = CENTER
                c.border = BORDER

        last_row = first_data_row + len(table.index) - 1
        last_col = get_column_letter(2 + len(table.columns))
        rng = f"C{first_data_row}:{last_col}{last_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="F8C9C4",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                end_type="max", end_color="ABEBC6",
            ),
        )

    def _build_price_data(self) -> None:
        ws = self.wb.create_sheet("Price History")
        ws.sheet_view.showGridLines = False
        df = self.data["Price Data"].tail(60).copy()

        keep = [c for c in ["Close", "Daily Return", "Cumulative Return",
                            "MA_20", "MA_50", "MA_200", "Signal"] if c in df.columns]
        df = df[keep]
        df.index = [d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)
                    for d in df.index]

        self._banner(ws, f"{self.ticker} — Price History (last 60 sessions)", "Close, returns, moving averages & trend signal", first="B", last=get_column_letter(2 + len(keep)))

        header = ["Date"] + keep
        hdr_row = 4
        self._table_header(ws, hdr_row, header, start_col=2)

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 13
        for i in range(len(keep)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14

        pct_cols = {"Daily Return", "Cumulative Return"}
        for i, (date, srow) in enumerate(df.iterrows()):
            r = hdr_row + 1 + i
            stripe = i % 2 == 1
            dcell = ws.cell(row=r, column=2, value=date)
            dcell.alignment = CENTER
            dcell.border = BORDER
            if stripe:
                dcell.fill = STRIPE_FILL
            for j, col in enumerate(keep):
                val = srow[col]
                c = ws.cell(row=r, column=3 + j)
                if col == "Signal":
                    c.value = val
                    c.alignment = CENTER
                    if val == "Bullish":
                        c.font = Font(bold=True, color=GREEN)
                    elif val == "Bearish":
                        c.font = Font(bold=True, color=RED)
                else:
                    c.value = None if pd.isna(val) else float(val)
                    c.number_format = PCT if col in pct_cols else MONEY2
                    c.alignment = RIGHT
                c.border = BORDER
                if stripe:
                    c.fill = STRIPE_FILL

        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=2)

    # -- styling helpers ------------------------------------------------------
    def _banner(self, ws: Worksheet, title: str, subtitle: str, first: str = "B", last: str = "C") -> None:
        ws.merge_cells(f"{first}1:{last}1")
        ws.merge_cells(f"{first}2:{last}2")
        t = ws[f"{first}1"]
        t.value = title
        t.font = TITLE_FONT
        t.fill = NAVY_FILL
        t.alignment = LEFT
        s = ws[f"{first}2"]
        s.value = subtitle
        s.font = SUB_FONT
        s.fill = NAVY_FILL
        s.alignment = LEFT
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 16

    def _section(self, ws: Worksheet, row: int, text: str, last_col: int = 3) -> int:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
        c = ws.cell(row=row, column=2, value=text)
        c.font = SECTION_FONT
        c.fill = LIGHT_FILL
        c.alignment = LEFT
        for col in range(2, last_col + 1):
            ws.cell(row=row, column=col).fill = LIGHT_FILL
        ws.row_dimensions[row].height = 20
        return row + 1

    def _kv(self, ws: Worksheet, row: int, label: str, value, num_fmt: str | None = None, prefix: str = "", color: str | None = None, emphasize: bool = False) -> int:
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = LABEL_FONT
        lc.alignment = LEFT
        lc.border = BORDER

        vc = ws.cell(row=row, column=3)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            vc.value = value
            if num_fmt:
                vc.number_format = (prefix + num_fmt) if prefix else num_fmt
        else:
            vc.value = f"{prefix}{value}" if value is not None else "N/A"
        font = Font(name="Calibri", size=12 if emphasize else 11,
                    bold=True, color=color or "1A1A1A")
        vc.font = font
        vc.alignment = RIGHT
        vc.border = BORDER
        return row + 1

    def _table_header(self, ws: Worksheet, row: int, headers: list[str], start_col: int = 2) -> None:
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=start_col + i, value=h)
            c.font = WHITE_BOLD
            c.fill = BLUE_FILL
            c.alignment = CENTER if i else LEFT
            c.border = BORDER
        ws.row_dimensions[row].height = 18

    def _table_row(self, ws: Worksheet, row: int, label: str, values: list, num_fmt: str, stripe: bool) -> int:
        fill = STRIPE_FILL if stripe else None
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = VALUE_FONT
        lc.alignment = LEFT
        lc.border = BORDER
        if fill:
            lc.fill = fill
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=3 + i, value=float(v))
            c.number_format = num_fmt
            c.alignment = RIGHT
            c.border = BORDER
            if fill:
                c.fill = fill
        return row + 1
